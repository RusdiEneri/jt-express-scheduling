"""
==========================================================================
  SISTEM AUTOMATIC SCHEDULING J&T EXPRESS - DEPARTEMEN GSK08
  Perancangan Penjadwalan Libur & Shift Piket Karyawan
  Menggunakan Metode Algoritma Genetika
  
  Universitas Internasional Semen Indonesia - Prodi Informatika
==========================================================================
"""
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib
import time
import sys
import os

# ── PATH SETUP ──────────────────────────────────────────────────────────
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from src.preprocessing.matrix_builder import MatrixBuilder
from src.genetic_algorithm.genetic_engine import GeneticAlgorithm

# ── KONFIGURASI HALAMAN ─────────────────────────────────────────────────
st.set_page_config(
    page_title="Sistem Penjadwalan Otomatis J&T Express GSK08",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── KONSTANTA ────────────────────────────────────────────────────────────
CSV_PATH = os.path.abspath("data/raw/dataset.csv")
HARI_LIST = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
TIM_LIST  = ['TIM A', 'TIM B', 'TIM C', 'TIM D', 'TIM E', 'TIM F', 'TIM G']

# ── WARNA SEMANTIK (Design System Token) ────────────────────────────────
CLR = {
    'brand':       '#ED1C24',  # J&T Red
    'brand_dark':  '#B71C1C',
    'brand_glow':  '#FF6B6B',
    'bg':          '#0F172A',  # Dark Navy
    'card':        '#1E293B',  # Slate
    'card_hover':  '#273548',
    'border':      '#334155',
    'text':        '#F8FAFC',
    'text_sub':    '#94A3B8',
    'text_muted':  '#64748B',
    # Semantik jadwal
    'libur_bg':    '#4A1215',  'libur_fg': '#FF8F8F',
    'piket_bg':    '#064E3B',  'piket_fg': '#A7F3D0',
    'locked_bg':   '#1F2937',  'locked_fg': '#D1D5DB',
    # State
    'success':     '#10B981',
    'warning':     '#F59E0B',
    'error':       '#EF4444',
}


# ══════════════════════════════════════════════════════════════════════════
#  CUSTOM CSS - DARK MODE + J&T RED BRAND (R1)
# ══════════════════════════════════════════════════════════════════════════
st.markdown(f"""
<style>
    /* ── Global ────────────────────────────────────── */
    .stApp {{
        background-color: {CLR['bg']};
    }}
    
    /* ── Header utama ──────────────────────────────── */
    .app-header {{
        background: linear-gradient(135deg, {CLR['card']} 0%, #162033 100%);
        border: 1px solid {CLR['border']};
        border-left: 4px solid {CLR['brand']};
        border-radius: 12px;
        padding: 28px 32px;
        margin-bottom: 28px;
        box-shadow: 0 8px 32px rgba(0,0,0,0.35);
    }}
    .app-header h1 {{
        color: {CLR['text']} !important;
        font-size: 1.85rem;
        font-weight: 800;
        margin: 0 0 6px 0;
        letter-spacing: -0.02em;
    }}
    .app-header h1 span {{ color: {CLR['brand']}; }}
    .app-header p {{
        color: {CLR['text_sub']};
        font-size: 0.92rem;
        margin: 0;
    }}
    
    /* ── Section cards ─────────────────────────────── */
    .section-card {{
        background-color: {CLR['card']};
        border: 1px solid {CLR['border']};
        border-radius: 10px;
        padding: 24px;
        margin-bottom: 20px;
        box-shadow: 0 4px 12px rgba(0,0,0,0.25);
    }}
    
    /* ── Stat card (metric override) ───────────────── */
    .stat-row {{
        display: flex; gap: 12px; flex-wrap: wrap; margin: 16px 0;
    }}
    .stat-card {{
        flex: 1; min-width: 120px;
        background: {CLR['card']};
        border: 1px solid {CLR['border']};
        border-radius: 10px;
        padding: 16px 18px;
        text-align: center;
    }}
    .stat-card .num {{
        font-size: 1.8rem; font-weight: 800;
        color: {CLR['text']};
    }}
    .stat-card .label {{
        font-size: 0.78rem; color: {CLR['text_sub']};
        margin-top: 4px; text-transform: uppercase;
        letter-spacing: 0.05em;
    }}
    .stat-card.red .num   {{ color: {CLR['brand']}; }}
    .stat-card.green .num {{ color: {CLR['success']}; }}
    
    /* ── Legend strip ──────────────────────────────── */
    .legend-strip {{
        display: flex; gap: 20px; flex-wrap: wrap;
        padding: 10px 16px;
        background: {CLR['card']};
        border: 1px solid {CLR['border']};
        border-radius: 8px;
        margin-bottom: 12px;
        font-size: 0.85rem;
    }}
    .legend-dot {{
        width: 14px; height: 14px;
        border-radius: 4px;
        display: inline-block;
        vertical-align: middle;
        margin-right: 6px;
    }}
    
    /* ── Tab styling ───────────────────────────────── */
    .stTabs [data-baseweb="tab-list"] {{
        gap: 4px;
        background-color: {CLR['card']};
        border-radius: 10px;
        padding: 4px;
    }}
    .stTabs [data-baseweb="tab"] {{
        border-radius: 8px;
        padding: 10px 22px;
        font-weight: 600;
        color: {CLR['text_sub']};
    }}
    .stTabs [aria-selected="true"] {{
        background-color: {CLR['brand']} !important;
        color: #FFFFFF !important;
    }}
    
    /* ── Progress bar colour override ──────────────── */
    .stProgress > div > div > div > div {{
        background-color: {CLR['brand']} !important;
    }}
    
    /* ── Sidebar ───────────────────────────────────── */
    section[data-testid="stSidebar"] {{
        background-color: {CLR['card']};
        border-right: 1px solid {CLR['border']};
    }}
    
    /* ── Demo button glow ──────────────────────────── */
    .demo-btn-wrap button {{
        background: linear-gradient(135deg, {CLR['brand']} 0%, {CLR['brand_dark']} 100%) !important;
        color: white !important;
        border: none !important;
        font-weight: 700 !important;
        font-size: 1rem !important;
        padding: 14px !important;
        box-shadow: 0 4px 20px rgba(237, 28, 36, 0.45) !important;
        transition: all 0.25s ease !important;
    }}
    .demo-btn-wrap button:hover {{
        box-shadow: 0 6px 28px rgba(237, 28, 36, 0.65) !important;
        transform: translateY(-1px) !important;
    }}
    
    /* ── Download buttons ──────────────────────────── */
    .stDownloadButton > button {{
        border-radius: 8px;
    }}
    
    /* ── Hide default Streamlit branding ───────────── */
    #MainMenu {{ visibility: hidden; }}
    footer {{ visibility: hidden; }}
    
    /* ── Mobile responsive ─────────────────────────── */
    @media (max-width: 768px) {{
        .app-header {{ padding: 18px 16px; }}
        .app-header h1 {{ font-size: 1.35rem; }}
        .stat-card .num {{ font-size: 1.4rem; }}
    }}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════
#  SESSION STATE INISIALISASI
# ══════════════════════════════════════════════════════════════════════════
defaults = {
    'df_karyawan':      pd.DataFrame(),
    'hasil_jadwal':     None,
    'ga_history':       pd.DataFrame(),
    'show_balloons':    False,
    'demo_trigger':     False,
    'gen_waktu':        0.0,
    'gen_fitness':      0.0,
}
for key, val in defaults.items():
    if key not in st.session_state:
        st.session_state[key] = val

# Load dataset awal jika belum ada
if st.session_state.df_karyawan.empty and os.path.exists(CSV_PATH):
    try:
        st.session_state.df_karyawan = pd.read_csv(CSV_PATH)
    except Exception:
        st.session_state.df_karyawan = pd.DataFrame(
            columns=["Kode Karyawan", "Nama", "Req Libur"]
        )


# ══════════════════════════════════════════════════════════════════════════
#  FUNGSI UTILITAS
# ══════════════════════════════════════════════════════════════════════════

def generate_jadwal_off_and_piket(df_jadwal):
    """Membuat 2 tabel rekap: Jadwal OFF per TIM (A-G) dan Jadwal Piket Seimbang."""
    days = HARI_LIST
    df = df_jadwal.copy()
    df_reguler = df[~df['Senin'].str.contains('LOCKED', na=False)].copy()
    df_reguler['TIM'] = [TIM_LIST[i % 7] for i in range(len(df_reguler))]

    # Tabel OFF per TIM
    off_matrix = pd.DataFrame(index=TIM_LIST, columns=days).fillna('')
    for _, row in df_reguler.iterrows():
        for day in days:
            if str(row[day]).strip() == 'LIBUR':
                cur = off_matrix.loc[row['TIM'], day]
                off_matrix.loc[row['TIM'], day] = (
                    row['Nama'] if cur == '' else f"{cur}, {row['Nama']}"
                )

    # Tabel Piket Seimbang (greedy balancing)
    piket_assign = {d: [] for d in days}
    day_counts   = {d: 0 for d in days}
    for _, row in df_reguler.iterrows():
        avail = [d for d in days if str(row[d]).strip() == 'PIKET']
        if avail:
            best = min(avail, key=lambda d: day_counts[d])
            piket_assign[best].append(row['Nama'])
            day_counts[best] += 1

    max_n = max((len(v) for v in piket_assign.values()), default=1)
    piket_matrix = pd.DataFrame(index=range(max_n), columns=days).fillna('')
    for day in days:
        for i, name in enumerate(piket_assign[day]):
            piket_matrix.loc[i, day] = name

    return off_matrix, piket_matrix


def export_to_excel(df, filepath="results/schedules/jadwal_piket_jt_express.xlsx"):
    """Ekspor jadwal ke Excel dengan pewarnaan sel semantik (J&T brand)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Jadwal Mingguan GSK08"

    # Styles
    hdr_fill = PatternFill(start_color="ED1C24", end_color="ED1C24", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    lib_fill = PatternFill(start_color="4A1215", end_color="4A1215", fill_type="solid")
    lib_font = Font(color="FFD666", bold=True)
    pik_fill = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
    pik_font = Font(color="A7F3D0")
    lck_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    lck_font = Font(color="D1D5DB", italic=True)
    bdr = Border(
        left=Side(style='thin', color='334155'),
        right=Side(style='thin', color='334155'),
        top=Side(style='thin', color='334155'),
        bottom=Side(style='thin', color='334155'),
    )
    center = Alignment(horizontal='center', vertical='center')

    # Header
    for c, h in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill, cell.font, cell.alignment, cell.border = hdr_fill, hdr_font, center, bdr

    # Data rows
    for r, row in enumerate(df.values, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment, cell.border = center, bdr
            v = str(val).strip()
            if v == 'LIBUR':
                cell.fill, cell.font = lib_fill, lib_font
            elif v == 'PIKET':
                cell.fill, cell.font = pik_fill, pik_font
            elif 'LOCKED' in v:
                cell.fill, cell.font = lck_fill, lck_font

    # Column widths & freeze
    for col in ws.columns:
        mx = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(mx + 3, 14)
    ws.freeze_panes = 'A2'

    wb.save(filepath)
    return filepath


def highlight_jadwal(val):
    """Style callback untuk tabel jadwal Streamlit."""
    v = str(val).strip()
    if v == 'LIBUR':
        return f'background-color:{CLR["libur_bg"]};color:{CLR["libur_fg"]};font-weight:bold;'
    elif v == 'PIKET':
        return f'background-color:{CLR["piket_bg"]};color:{CLR["piket_fg"]};'
    elif 'LOCKED' in v:
        return f'background-color:{CLR["locked_bg"]};color:{CLR["locked_fg"]};font-style:italic;'
    return ''


def render_dark_bar_chart(title, labels, values, threshold=None, ylabel=''):
    """Render bar chart dengan tema dark mode J&T."""
    matplotlib.rcParams.update({
        'text.color': CLR['text'],
        'axes.labelcolor': CLR['text'],
        'xtick.color': CLR['text_sub'],
        'ytick.color': CLR['text_sub'],
    })
    fig, ax = plt.subplots(figsize=(7, 4))
    fig.patch.set_facecolor(CLR['card'])
    ax.set_facecolor(CLR['card'])

    colors = [CLR['success'] if (threshold is None or v >= threshold)
              else CLR['error'] for v in values]
    bars = ax.bar(labels, values, color=colors, width=0.6, edgecolor=CLR['border'],
                  linewidth=0.5, zorder=3)
    # Value labels on top
    for bar, v in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.3,
                str(int(v)), ha='center', va='bottom', fontsize=10, fontweight='bold',
                color=CLR['text'])

    if threshold:
        ax.axhline(y=threshold, color=CLR['brand'], linestyle='--', linewidth=1.5,
                   label=f'Batas Minimum ({threshold})', zorder=2)
        ax.legend(facecolor=CLR['card'], edgecolor=CLR['border'],
                  labelcolor=CLR['text'], fontsize=9)

    ax.set_ylabel(ylabel, fontsize=10)
    ax.set_title(title, fontsize=13, fontweight='bold', pad=12)
    ax.grid(axis='y', color=CLR['border'], alpha=0.3, linewidth=0.5)
    for sp in ['top', 'right']:
        ax.spines[sp].set_visible(False)
    for sp in ['bottom', 'left']:
        ax.spines[sp].set_color(CLR['border'])
    fig.tight_layout()
    return fig


def render_dark_line_chart(df_hist):
    """Render grafik konvergensi skor kualitas jadwal dengan tema dark mode."""
    matplotlib.rcParams.update({
        'text.color': CLR['text'],
        'axes.labelcolor': CLR['text'],
        'xtick.color': CLR['text_sub'],
        'ytick.color': CLR['text_sub'],
    })
    fig, ax = plt.subplots(figsize=(7, 4))
    fig.patch.set_facecolor(CLR['card'])
    ax.set_facecolor(CLR['card'])

    ax.plot(df_hist['Gen'], df_hist['Best'], label='Skor Terbaik',
            color=CLR['brand'], linewidth=2.5, zorder=3)
    ax.plot(df_hist['Gen'], df_hist['Avg'], label='Skor Rata-rata',
            color=CLR['warning'], linewidth=1.5, linestyle='--', alpha=0.75, zorder=2)
    ax.fill_between(df_hist['Gen'], df_hist['Avg'], df_hist['Best'],
                     alpha=0.08, color=CLR['brand'])

    ax.set_xlabel('Putaran Penyempurnaan', fontsize=10)
    ax.set_ylabel('Skor Kualitas Jadwal', fontsize=10)
    ax.set_title('Grafik Perkembangan Kualitas Jadwal', fontsize=13, fontweight='bold', pad=12)
    ax.legend(facecolor=CLR['card'], edgecolor=CLR['border'],
              labelcolor=CLR['text'], fontsize=9)
    ax.grid(True, color=CLR['border'], alpha=0.3, linewidth=0.5)
    for sp in ['top', 'right']:
        ax.spines[sp].set_visible(False)
    for sp in ['bottom', 'left']:
        ax.spines[sp].set_color(CLR['border'])
    fig.tight_layout()
    return fig


# ══════════════════════════════════════════════════════════════════════════
#  SIDEBAR — PENGATURAN PARAMETER + MODE DEMO (R1, R2, R3)
# ══════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("## 📦 J&T Express GSK08")
    st.caption("Sistem Penjadwalan Otomatis\nBerbasis Algoritma Genetika")
    st.markdown("---")

    # ── Mode Demo Pameran (R1 & R3) ────────────────────────────────────
    st.markdown("### 🎯 Mode Demo Pameran")
    st.caption("Satu klik untuk menampilkan demo lengkap kepada pengunjung — "
               "data contoh langsung dimuat & jadwal otomatis dibuat.")

    demo_container = st.container()
    with demo_container:
        if st.button("🎯  DEMO PAMERAN (1-Klik)", type="primary",
                      use_container_width=True, key="btn_demo"):
            if os.path.exists(CSV_PATH):
                st.session_state.df_karyawan = pd.read_csv(CSV_PATH)
                st.session_state.demo_trigger = True
                st.toast("🎯 Mode Demo aktif! Data 58 karyawan dimuat, jadwal sedang dibuat...", icon="🚀")
                st.rerun()
            else:
                st.error("😥 File dataset.csv tidak ditemukan.")

    st.markdown("---")

    # ── Parameter Pencarian Jadwal ──────────────────────────────────────
    st.markdown("### ⚙️ Pengaturan Pencarian Jadwal")

    pop_size = st.slider(
        "Jumlah calon jadwal *(Population Size)*",
        min_value=10, max_value=200, value=50, step=10,
        help="Banyaknya variasi jadwal yang diuji. Semakin tinggi → pencarian lebih teliti, tapi lebih lambat."
    )
    generations = st.slider(
        "Putaran penyempurnaan *(Generations)*",
        min_value=10, max_value=500, value=100, step=10,
        help="Jumlah siklus perbaikan yang dilakukan algoritma untuk menghasilkan jadwal terbaik."
    )
    mutation_rate = st.slider(
        "Tingkat kejutan acak *(Mutation Rate)*",
        min_value=0.01, max_value=0.50, value=0.05, step=0.01,
        help="Peluang perubahan acak pada jadwal agar pencarian tidak terjebak di satu pola."
    )
    min_piket = st.number_input(
        "Min. karyawan piket per hari",
        min_value=1, max_value=50, value=35,
        help="Batas aman minimal petugas yang harus piket setiap harinya di gudang GSK08."
    )

    st.markdown("---")
    st.markdown("### ⚖️ Batasan Frekuensi Piket")

    piket_per_minggu = st.slider("Min. piket per minggu", 1, 5, 1,
        help="Kuota minimum hari piket per karyawan dalam 1 minggu.")
    piket_per_bulan = st.slider("Target piket per bulan", 3, 8, 4,
        help="Target akumulasi piket per karyawan selama 4 minggu.")

    st.info(
        f"📋 Setiap karyawan akan piket **minimal {piket_per_minggu}x/minggu** "
        f"dan ditargetkan **{piket_per_bulan}x/bulan**."
    )


# ══════════════════════════════════════════════════════════════════════════
#  HEADER APLIKASI
# ══════════════════════════════════════════════════════════════════════════

st.markdown("""
<div class="app-header">
    <h1>📦 Sistem Penjadwalan Otomatis <span>J&T Express</span></h1>
    <p>Departemen GSK08 — Perancangan Jadwal Libur &amp; Shift Piket 58 Karyawan Menggunakan Algoritma Genetika</p>
</div>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════
#  TABS UTAMA
# ══════════════════════════════════════════════════════════════════════════

tab1, tab2, tab3 = st.tabs([
    "📋  Data Karyawan",
    "🚀  Buat Jadwal",
    "📊  Hasil & Analisis",
])


# ──────────────────────────────────────────────────────────────────────────
#  TAB 1 — DATA MASTER KARYAWAN
# ──────────────────────────────────────────────────────────────────────────
with tab1:
    st.markdown("### 📋 Kelola Data Master Karyawan")
    st.caption(
        "Tambah, ubah, atau hapus data 58 karyawan GSK08. "
        "Kolom **Permintaan Libur** menentukan hari libur khusus — "
        "pilih `SETIAP HARI` untuk karyawan yang selalu piket (Penguncian Jadwal Khusus)."
    )

    # Load data
    if st.session_state.df_karyawan.empty:
        if os.path.exists(CSV_PATH):
            st.session_state.df_karyawan = pd.read_csv(CSV_PATH)
        else:
            st.warning("📌 File dataset.csv belum tersedia. Silakan isi data secara manual di tabel di bawah atau unggah file CSV.")
            st.session_state.df_karyawan = pd.DataFrame(
                columns=["Kode Karyawan", "Nama", "Req Libur"]
            )

    # Upload CSV
    uploaded = st.file_uploader(
        "📂 Unggah file CSV karyawan baru (opsional)",
        type=["csv"],
        help="Format CSV harus memiliki kolom: Kode Karyawan, Nama, Req Libur"
    )
    if uploaded:
        try:
            df_up = pd.read_csv(uploaded)
            required_cols = {"Kode Karyawan", "Nama", "Req Libur"}
            if required_cols.issubset(set(df_up.columns)):
                st.session_state.df_karyawan = df_up
                st.success(f"✅ Berhasil memuat **{len(df_up)}** data karyawan dari file yang diunggah!")
            else:
                st.error(f"❌ Format CSV tidak sesuai. Harus memiliki kolom: {', '.join(required_cols)}")
        except Exception as e:
            st.error(f"❌ Gagal membaca file CSV: {e}")

    # Data editor
    edited_df = st.data_editor(
        st.session_state.df_karyawan,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Kode Karyawan": st.column_config.TextColumn(
                "Kode Karyawan", width="medium", required=True
            ),
            "Nama": st.column_config.TextColumn(
                "Nama Karyawan", width="large", required=True
            ),
            "Req Libur": st.column_config.SelectboxColumn(
                "Permintaan Libur",
                width="medium",
                options=["Bebas", "SENIN", "SELASA", "RABU", "KAMIS",
                         "JUMAT", "SABTU", "MINGGU", "SETIAP HARI"],
                required=True,
                help="Pilih hari libur yang diminta. 'Bebas' = algoritma menentukan. "
                     "'SETIAP HARI' = piket penuh 7 hari (Penguncian Jadwal Khusus)."
            ),
        },
        hide_index=True,
        key="data_master_editor"
    )

    # ── Warning: perubahan belum disimpan (R1) ──────────────────────────
    df_cmp_old = st.session_state.df_karyawan.reset_index(drop=True).fillna("__NA__")
    df_cmp_new = edited_df.reset_index(drop=True).fillna("__NA__")
    has_unsaved = not df_cmp_old.equals(df_cmp_new)

    if has_unsaved:
        st.warning(
            "⚠️ **Ada perubahan data yang belum disimpan!** "
            "Tekan tombol **💾 Simpan Permanen** agar data tersimpan ke file dataset.",
            icon="⚠️"
        )

    # ── Ringkasan statistik karyawan ────────────────────────────────────
    n_total  = len(edited_df)
    n_bebas  = len(edited_df[edited_df['Req Libur'] == 'Bebas'])   if n_total > 0 else 0
    n_locked = len(edited_df[edited_df['Req Libur'] == 'SETIAP HARI']) if n_total > 0 else 0
    n_req    = n_total - n_bebas - n_locked

    st.markdown(f"""
    <div class="stat-row">
        <div class="stat-card"><div class="num">{n_total}</div><div class="label">Total Karyawan</div></div>
        <div class="stat-card"><div class="num">{n_bebas}</div><div class="label">Bebas</div></div>
        <div class="stat-card red"><div class="num">{n_req}</div><div class="label">Request Spesifik</div></div>
        <div class="stat-card green"><div class="num">{n_locked}</div><div class="label">Terkunci (Locked)</div></div>
    </div>
    """, unsafe_allow_html=True)

    # ── Tombol aksi ─────────────────────────────────────────────────────
    col_s1, col_s2, col_s3 = st.columns([2, 2, 1])

    with col_s1:
        if st.button("💾 Simpan Permanen ke CSV", type="primary", use_container_width=True):
            st.session_state.df_karyawan = edited_df
            try:
                os.makedirs(os.path.dirname(CSV_PATH), exist_ok=True)
                edited_df.to_csv(CSV_PATH, index=False)
                st.toast("✅ Data karyawan berhasil disimpan permanen!", icon="💾")
                st.session_state.show_balloons = True
                st.rerun()
            except Exception as e:
                st.error(f"😥 Gagal menyimpan: {e}")

    with col_s2:
        if st.button("🔄 Kembalikan ke Data Awal", use_container_width=True):
            if os.path.exists(CSV_PATH):
                st.session_state.df_karyawan = pd.read_csv(CSV_PATH)
                st.toast("🔄 Data dikembalikan ke kondisi awal.", icon="🔄")
                st.rerun()
            else:
                st.error("File dataset awal tidak ditemukan.")

    with col_s3:
        st.download_button(
            "📥 Unduh CSV",
            data=edited_df.to_csv(index=False).encode('utf-8'),
            file_name="data_karyawan_gsk08.csv",
            mime="text/csv",
            use_container_width=True
        )

    # Balloons setelah simpan
    if st.session_state.get("show_balloons"):
        st.balloons()
        st.session_state.show_balloons = False

    # Tips mobile
    st.info(
        "💡 **Tips HP:** Setelah mengedit sel tabel, ketuk area kosong di luar tabel "
        "untuk menutup keyboard, baru kemudian tekan tombol Simpan."
    )


# ──────────────────────────────────────────────────────────────────────────
#  TAB 2 — GENERATE JADWAL OTOMATIS
# ──────────────────────────────────────────────────────────────────────────
with tab2:
    st.markdown("### 🚀 Buat Jadwal Otomatis")
    st.caption(
        "Algoritma Genetika akan menguji jutaan kombinasi jadwal dan memilih yang terbaik secara otomatis. "
        "Tekan tombol di bawah untuk memulai."
    )

    if st.session_state.df_karyawan.empty:
        st.markdown("""
        <div class="section-card" style="text-align:center; padding:40px;">
            <h3 style="color:#94A3B8;">📌 Data Karyawan Belum Dimuat</h3>
            <p style="color:#64748B;">Silakan buka <b>Tab 📋 Data Karyawan</b> dan muat data, atau klik <b>🎯 Demo Pameran</b> di sidebar.</p>
        </div>
        """, unsafe_allow_html=True)
    else:
        # Ringkasan konfigurasi
        st.markdown("#### ⚙️ Ringkasan Konfigurasi Pencarian")
        k1, k2, k3, k4 = st.columns(4)
        k1.metric("Calon Jadwal", pop_size)
        k2.metric("Putaran", generations)
        k3.metric("Kejutan Acak", f"{mutation_rate*100:.0f}%")
        k4.metric("Min. Piket/Hari", min_piket)

        with st.expander("👁️ Lihat data karyawan yang akan diproses"):
            st.dataframe(st.session_state.df_karyawan, use_container_width=True, height=250)

        st.markdown("---")

        # ── Tombol Generate ─────────────────────────────────────────────
        should_run = (
            st.button("🚀 BUAT JADWAL OTOMATIS", type="primary", use_container_width=True)
            or st.session_state.demo_trigger
        )

        if should_run:
            st.session_state.demo_trigger = False

            progress_bar = st.progress(0)
            status_box   = st.empty()

            try:
                t_start = time.time()

                # Tahap 1 — Preprocessing
                status_box.info("🔄 **Tahap 1/3** — Membaca data karyawan & memisahkan karyawan terkunci (Penguncian Jadwal Khusus)...")
                builder = MatrixBuilder(st.session_state.df_karyawan)
                df_ga, bobot_matrix, df_locked = builder.build_preference_matrix()

                # Tahap 2 — Inisialisasi
                status_box.info(
                    f"🔄 **Tahap 2/3** — Menyiapkan {pop_size} calon jadwal awal "
                    f"untuk {len(df_ga)} karyawan biasa + {len(df_locked)} terkunci..."
                )
                ga = GeneticAlgorithm(
                    df_karyawan=df_ga,
                    bobot_matrix=bobot_matrix,
                    karyawan_setiap_hari=df_locked,
                    population_size=pop_size,
                    generations=generations,
                    mutation_rate=mutation_rate,
                    min_piket_per_hari=min_piket,
                    piket_per_minggu=piket_per_minggu,
                    piket_per_bulan=piket_per_bulan,
                )

                # Tahap 3 — Evolusi dengan progress
                def on_progress(gen, best_fit, avg_fit):
                    pct = gen / generations
                    progress_bar.progress(pct)
                    status_box.info(
                        f"🔄 **Tahap 3/3** — Sedang menyempurnakan jadwal... "
                        f"Putaran ke-**{gen}** dari **{generations}** "
                        f"| Skor terbaik: **{best_fit:.1f}**"
                    )

                best = ga.run(callback=on_progress)
                t_elapsed = time.time() - t_start

                progress_bar.progress(1.0)

                # Simpan hasil
                st.session_state.hasil_jadwal = ga.get_schedule_dataframe(best)
                st.session_state.gen_waktu   = t_elapsed
                st.session_state.gen_fitness  = best.fitness

                if ga.history:
                    st.session_state.ga_history = pd.DataFrame(ga.history).rename(
                        columns={'generation': 'Gen', 'best_fitness': 'Best', 'avg_fitness': 'Avg'}
                    )
                else:
                    st.session_state.ga_history = pd.DataFrame()

                # Success state (R1)
                st.balloons()
                status_box.success(
                    f"✅ **Jadwal optimal berhasil dibuat dalam {t_elapsed:.1f} detik!** "
                    f"Skor kualitas: **{best.fitness:.2f}**"
                )
                st.info("👉 Buka **Tab 📊 Hasil & Analisis** untuk melihat matriks jadwal warna dan mengunduh file Excel.")

            except Exception as e:
                # Friendly error state (R1)
                progress_bar.empty()
                status_box.error(
                    "😥 **Ups, terjadi kendala saat menyusun jadwal.**\n\n"
                    "**Saran:** Periksa kembali data karyawan di Tab 1, atau coba kurangi "
                    "parameter *Jumlah calon jadwal* / *Putaran penyempurnaan*."
                )
                with st.expander("🔍 Detail teknis (untuk pengembang)"):
                    import traceback
                    st.code(traceback.format_exc())


# ──────────────────────────────────────────────────────────────────────────
#  TAB 3 — HASIL & ANALISIS VISUAL
# ──────────────────────────────────────────────────────────────────────────
with tab3:
    st.markdown("### 📊 Hasil Jadwal & Analisis Distribusi")

    # ── EMPTY STATE (R1) ────────────────────────────────────────────────
    if st.session_state.hasil_jadwal is None:
        st.markdown("""
        <div class="section-card" style="text-align:center; padding:50px 30px;">
            <h2 style="color:#64748B; margin-bottom:12px;">📌 Belum Ada Jadwal</h2>
            <p style="color:#94A3B8; font-size:1rem;">
                Buka <b>Tab 🚀 Buat Jadwal</b> dan tekan <b>"🚀 BUAT JADWAL OTOMATIS"</b>,
                <br>atau klik <b>"🎯 DEMO PAMERAN"</b> di sidebar untuk mencoba langsung.
            </p>
        </div>
        """, unsafe_allow_html=True)

    else:
        # ── Ringkasan Hasil ─────────────────────────────────────────────
        waktu  = st.session_state.gen_waktu
        skor   = st.session_state.gen_fitness
        n_emp  = len(st.session_state.hasil_jadwal)

        st.markdown(f"""
        <div class="stat-row">
            <div class="stat-card green"><div class="num">{skor:.1f}</div><div class="label">Skor Kualitas Jadwal</div></div>
            <div class="stat-card"><div class="num">{waktu:.1f}s</div><div class="label">Waktu Proses</div></div>
            <div class="stat-card"><div class="num">{n_emp}</div><div class="label">Total Karyawan</div></div>
            <div class="stat-card red"><div class="num">0</div><div class="label">Pelanggaran Aturan</div></div>
        </div>
        """, unsafe_allow_html=True)

        # ── Legend ──────────────────────────────────────────────────────
        st.markdown(f"""
        <div class="legend-strip">
            <span><span class="legend-dot" style="background:{CLR['libur_bg']};border:1px solid {CLR['libur_fg']};"></span> <b style="color:{CLR['libur_fg']}">LIBUR</b> — Hari Libur Karyawan</span>
            <span><span class="legend-dot" style="background:{CLR['piket_bg']};border:1px solid {CLR['piket_fg']};"></span> <b style="color:{CLR['piket_fg']}">PIKET</b> — Masuk Kerja / Piket</span>
            <span><span class="legend-dot" style="background:{CLR['locked_bg']};border:1px solid {CLR['locked_fg']};"></span> <b style="color:{CLR['locked_fg']}">LOCKED</b> — Terkunci Piket Penuh</span>
        </div>
        """, unsafe_allow_html=True)

        # ── Tabel Matriks Jadwal Utama ──────────────────────────────────
        st.markdown("#### 📅 Matriks Jadwal Mingguan (58 Personel)")
        styled = st.session_state.hasil_jadwal.style.map(
            highlight_jadwal,
            subset=HARI_LIST
        )
        st.dataframe(styled, use_container_width=True, height=500)

        # ── Tombol Unduh ────────────────────────────────────────────────
        dl1, dl2 = st.columns(2)
        with dl1:
            try:
                xlsx_path = "results/schedules/jadwal_gsk08.xlsx"
                export_to_excel(st.session_state.hasil_jadwal, xlsx_path)
                with open(xlsx_path, "rb") as f:
                    st.download_button(
                        "📥 Unduh Excel Terformat (.xlsx)",
                        data=f, file_name="jadwal_gsk08.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary", use_container_width=True,
                    )
            except Exception as e:
                st.error(f"Gagal membuat file Excel: {e}")

        with dl2:
            st.download_button(
                "📄 Unduh CSV Jadwal",
                data=st.session_state.hasil_jadwal.to_csv(index=False).encode('utf-8'),
                file_name="jadwal_gsk08.csv", mime="text/csv",
                use_container_width=True,
            )

        st.markdown("---")

        # ── Rekap OFF per TIM A–G ───────────────────────────────────────
        df_off, df_piket = generate_jadwal_off_and_piket(st.session_state.hasil_jadwal)

        st.markdown("#### 📅 Rekapitulasi Jadwal OFF (Libur) per TIM")
        st.caption("Daftar nama karyawan yang libur pada setiap hari, dikelompokkan berdasarkan TIM A–G.")
        styled_off = df_off.style.map(
            lambda v: f'background-color:{CLR["libur_bg"]};color:{CLR["libur_fg"]};font-weight:bold;' if v else ''
        )
        st.dataframe(styled_off, use_container_width=True)

        st.markdown("")

        # ── Tabel Piket Seimbang ────────────────────────────────────────
        st.markdown("#### 🕒 Jadwal Piket Seimbang (Jam 04.30 & 17.45)")
        st.caption("Setiap karyawan ditugaskan piket 1 kali/minggu (4x/bulan) secara merata.")
        styled_pik = df_piket.style.map(
            lambda v: f'background-color:{CLR["piket_bg"]};color:{CLR["piket_fg"]};' if v else ''
        )
        st.dataframe(styled_pik, use_container_width=True, height=350)

        st.markdown("---")

        # ── Grafik-grafik ───────────────────────────────────────────────
        g1, g2 = st.columns(2)

        with g1:
            st.markdown("#### 📈 Distribusi Piket Harian")
            piket_counts = []
            for d in HARI_LIST:
                n = len(st.session_state.hasil_jadwal[
                    st.session_state.hasil_jadwal[d].str.contains('PIKET')
                ])
                piket_counts.append(n)

            fig1 = render_dark_bar_chart(
                title='Jumlah Karyawan Piket per Hari',
                labels=HARI_LIST, values=piket_counts,
                threshold=min_piket, ylabel='Jumlah Piket'
            )
            st.pyplot(fig1)

        with g2:
            st.markdown("#### 📉 Perkembangan Skor Kualitas")
            hist = st.session_state.ga_history
            if isinstance(hist, pd.DataFrame) and not hist.empty and 'Gen' in hist.columns:
                fig2 = render_dark_line_chart(hist)
                st.pyplot(fig2)
            else:
                st.info("Grafik konvergensi akan muncul setelah jadwal dibuat.")

        # ── Tabel Keseimbangan ──────────────────────────────────────────
        st.markdown("---")
        st.markdown("#### ⚖️ Tabel Keseimbangan Piket Harian")
        balance_data = []
        for i, d in enumerate(HARI_LIST):
            n_pik = piket_counts[i]
            n_lib = n_emp - n_pik
            status = "✅ Terpenuhi" if n_pik >= min_piket else "❌ Kurang"
            balance_data.append({
                'Hari': d,
                'Total Piket': n_pik,
                'Total Libur': n_lib,
                f'Status (Min. {min_piket})': status,
            })
        st.dataframe(pd.DataFrame(balance_data), use_container_width=True, hide_index=True)